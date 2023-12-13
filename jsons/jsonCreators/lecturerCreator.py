import json
from openpyxl import Workbook,load_workbook
Tr2Eng = str.maketrans("çğıöşüÖÜİ", "cgiosuOUI")
# Load the workbook
wb = load_workbook(filename='data.xlsx')
ws = wb['Page1']

# Create the data dictionary

data = {}

for i in range(2, 148): 
    lect = ws.cell(row=i, column=3).value
    data[lect] =[]
    for j in range(2, 148):
        if lect == ws.cell(row=j, column=3).value:
            data[lect].append(ws.cell(row=j, column=1).value)

    


lecturers = []
id = 150102

for(key, value) in data.items():
   
     # input_string[0] + input_string[1:].lower()
    split = key.split(" ")
    
    if(len(split) > 3):
        name = split[1] + " " + split[2]
        surname = split[3]
    else:
        name = split[1]
        surname = split[2]
    
    name = name[0] + name[1:].lower()
    surname = surname[0] + surname[1:].lower()
    split[1] = split[1][0] + split[1][1:].lower()
    
    id += 1
    password = (str)(id) + split[1]
    # print(name, surname, id, password)
    title = split[0]

    name = name.translate(Tr2Eng)
    surname = surname.translate(Tr2Eng)
    title = title.translate(Tr2Eng)
    password = password.translate(Tr2Eng)


    courses = value
    # print(value)

    lecturer = {
        "Name": name,
        "Surname": surname,
        "EMail": split[1]+surname+"@marmara.edu.com",
        "PhoneNumber": "05555555550",
        "ID": (str)(id),
        "Password": password,
        "Faculty": "Engineering",
        "Department": "Computer Science Engineering",
        "Courses": courses,
        "AcademicTitle": title
    }

    lecturers.append(lecturer)



with open("lecturers.json", "w", encoding="utf-8") as outfile:
    json.dump(lecturers, outfile, ensure_ascii=False)
